"""
Creates a manual labeling Excel workbook for AI-CFT assessment.
Run: python3 create_labeling_excel.py
Output: AI_CFT_Labeling.xlsx
"""

import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Color palette ──────────────────────────────────────────────────────────────
CLR = {
    "header_dark":  "1F3864",   # dark navy
    "header_mid":   "2E75B6",   # medium blue
    "header_light": "BDD7EE",   # pale blue
    "acquire":      "E2EFDA",   # light green
    "deepen":       "FFF2CC",   # light yellow
    "create":       "FCE4D6",   # light orange
    "note":         "F2F2F2",   # light grey
    "required":     "FF0000",   # red (asterisk)
    "white":        "FFFFFF",
    "row_alt":      "EAF4FB",
}

FONT_WHITE_BOLD = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
FONT_DARK_BOLD  = Font(name="Calibri", bold=True, color="1F3864", size=10)
FONT_NORMAL     = Font(name="Calibri", size=10)
FONT_SMALL      = Font(name="Calibri", size=9)
FONT_TINY       = Font(name="Calibri", size=8, color="595959")

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)

def thin_border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

# ── Student lists ──────────────────────────────────────────────────────────────
STUDENTS_2025 = [
    "Ozzy", "Ally", "Bella", "Bob", "Daisy",
    "Daniella", "David", "Eliot", "Henry", "Irene",
    "Isabel", "Karl", "Michael", "Mike", "Nicolas",
    "Zabby", "Adam", "Eddy", "Aden", "Barbara",
    "Boris", "Calvin", "Darby", "Demi", "Daryl",
    "Edgar", "Frank", "Felicity", "Kim", "Sabrina",
]

# 2026: class-level PDFs — add student IDs / pseudonyms as you learn them
STUDENTS_2026 = [
    "Sheila", "Marco", "Iris", "Marcus", "Isabel",
    "Amy", "Serena", "Irma", "Kate", "Bruno",
    "Zara", "Shana", "Melinda", "Helena", "Nadia",
    "Ulysses",
]

# ── Worksheet task definitions ─────────────────────────────────────────────────
# Each entry: (column_header, description_shown_at_bottom, ai_cft_level)
# column_header = exact question/blank ID as it appears in the worksheet
# description   = what the student must do / what correct answer looks like
# ai_cft_level  = Acquire / Deepen / Create / "" (for Overall / Notes cols)

WS_TASKS = {

    # ── WS1: Fill-in-the-blank — Important Terms ───────────────────────────────
    # 11 numbered blanks as they appear on the worksheet (red numbers in the answer key)
    "WS1 - Terms": [
        ("Blank 1\nLabel name",
         "Diagram arrow pointing to food card name -> 'label' (etiket)",
         "Acquire"),
        ("Blank 2\nObject name",
         "Diagram arrow pointing to 'Findikli Gofret' card -> 'object' (nesne)",
         "Acquire"),
        ("Blank 3\nFeature name",
         "Diagram arrow pointing to left column of card -> 'feature / variable / characteristic'",
         "Acquire"),
        ("Blank 4\nFeature value",
         "Diagram arrow pointing to a numeric cell -> 'value of the feature / characteristic'",
         "Acquire"),
        ("Blank 5\nObject definition",
         "Cloze: 'Individual foods are called ___' -> object (nesne)",
         "Acquire"),
        ("Blank 6\nObject example",
         "Cloze: 'Hazelnut wafer is a ___' -> object (nesne)",
         "Acquire"),
        ("Blank 7\nFeature term",
         "Cloze: left-column names are called '___ / characteristic / variable'",
         "Acquire"),
        ("Blank 8\nFeature count",
         "Cloze: 'How many features? ___' -> 7 (yedi / seven)",
         "Acquire"),
        ("Blank 9\nFeature list",
         "Cloze: student writes all feature names: Energy, Fat, Saturated Fat, Carbohydrates, Sugar, Protein, Salt",
         "Acquire"),
        ("Blank 10\nObject example 2",
         "Cloze: 'For example ___ contains 31.9g fat' -> Hazelnut Wafer (Findikli Gofret)",
         "Acquire"),
        ("Blank 11\nLabel term",
         "Cloze: 'Hazelnut wafer has been given ___ as label' -> not recommended (tavsiye edilemez)",
         "Acquire"),
        ("Overall WS1", "Overall AI-CFT level for WS1 as a whole", ""),
        ("Notes WS1", "Free notes: edge cases, illegible answers, etc.", ""),
    ],

    # ── WS3: Applying Thresholds in a Decision Rule ────────────────────────────
    # Worksheet has: Task 1 (popcorn), Task 2 (apple), Task 3 (fries), Task 4 (own threshold)
    # Each task contains a classification blank AND a justification blank
    "WS3 - Threshold Apply": [
        ("Task 1\nPopcorn: class + reason",
         "Classifies popcorn as recommended (fat <= 8g) AND gives correct justification",
         "Acquire"),
        ("Task 2\nApple: class + reason",
         "Classifies apple as recommended AND gives correct justification referencing fat value",
         "Acquire"),
        ("Task 3\nFries: class + reason",
         "Classifies french fries as not recommended (fat > 8g) AND gives correct justification",
         "Acquire"),
        ("Task 4\nOwn energy threshold",
         "Writes own threshold for the Energy variable with correct <= / > notation in the diagram",
         "Deepen"),
        ("Overall WS3", "Overall AI-CFT level for WS3 as a whole", ""),
        ("Notes WS3", "Free notes", ""),
    ],

    # ── WS4: Searching for the Best Threshold Value ────────────────────────────
    # Pipeline blanks B1–B5 (aligned with rubrics/WS4_rubric.json)
    "WS4 - Best Threshold": [
        ("B1\nAvocado–fries threshold",
         "Blank 1: Draws or states new threshold between avocado and french fries (yeni eşik)",
         "Deepen"),
        ("B2\nFour misclassified foods",
         "Blank 2: Marks or writes all four foods — jelibon, kraker, yulaf, avokado (order irrelevant; 3 = fail)",
         "Deepen"),
        ("B3\nFewer misclassifications",
         "Blank 3: States misclassification decreased after the improved threshold",
         "Deepen"),
        ("B4\nPia / same fat",
         "Blank 4: Agrees Pia is haklı because apple and raspberry jam have the same fat value (not meta peer commentary)",
         "Deepen"),
        ("B5\nEnergy threshold",
         "Blank 5: Writes learned energy threshold in range 160–2223",
         "Deepen"),
        ("Overall WS4", "Overall AI-CFT level for WS4 as a whole", ""),
        ("Notes WS4", "Free notes", ""),
    ],

    # ── WS5: Trying Out Threshold Values ───────────────────────────────────────
    # Grid rows validated against data/prodabi_food_cards.csv (N=11)
    "WS5 - Try Thresholds": [
        ("Row 1\nThreshold + counts",
         "Row 1: Valid feature + inclusive ≤/≥ operator; correct+errors=11; MCR=errors/11; counts match food cards",
         "Deepen"),
        ("Row 2\nThreshold + counts",
         "Row 2: Same rules; different feature/threshold trial",
         "Deepen"),
        ("Row 3\nThreshold + counts",
         "Row 3: Same rules",
         "Deepen"),
        ("Operator check",
         "Strict < or > only (no ≤/≥) → partial at most — equality value unclassified",
         "Deepen"),
        ("B25\nFinal threshold",
         "B25: chooses threshold with lowest yanlış sınıflandırma count among valid grid rows; flag tie when multiple minima",
         "Deepen"),
        ("Overall WS5", "Overall AI-CFT level for WS5 as a whole", ""),
        ("Notes WS5", "Free notes", ""),
    ],

    # ── WS6: Two-Level Decision Tree (11 food cards, same as WS5) ─────────────
    "WS6 - Draw Tree": [
        ("Root\nFeature + threshold",
         "B1/B2: valid feature + inclusive ≤/≥ on B2 (strict < or > → partial, same as WS5)",
         "Deepen"),
        ("Root\nBranch labels",
         "B3/B4: evet (≤) and hayır (>) branch labels",
         "Deepen"),
        ("Inner node\nFeature + threshold",
         "B6/B7: second feature (≠ root) + threshold — çift seviyeli",
         "Deepen"),
        ("Inner\nBranch labels",
         "B8/B9: inner evet/hayır labels",
         "Deepen"),
        ("Leaves\nClass labels",
         "B10/B11/B13: tavsiye edilir/edilmez on required paths",
         "Deepen"),
        ("Tree validity",
         "Two levels, 2 features, paired operators; MCR=0 with 2 levels is OK",
         "Deepen"),
        ("Overall WS6", "Overall AI-CFT level for WS6 as a whole", ""),
        ("Notes WS6", "Free notes", ""),
    ],

    # ── WS7: Formulate Decision Rules ─────────────────────────────────────────
    # Page 1: matching task (paths A, B, C). Page 2: path count blank + Rules 1-6
    "WS7 - Rules": [
        ("Page 1\nPath A match",
         "Page 1 matching: assigns correct rule to Path A (energy < 180 kcal -> recommended)",
         "Acquire"),
        ("Page 1\nPath B match",
         "Page 1 matching: assigns correct rule to Path B (energy >= 180 AND protein < 7.7 -> not recommended)",
         "Acquire"),
        ("Page 1\nPath C match",
         "Page 1 matching: assigns correct rule to Path C (energy >= 180 AND protein >= 7.7 -> recommended)",
         "Acquire"),
        ("Page 2\nPath count",
         "Page 2 blank: 'My tree contains ___ paths' — states correct number of paths",
         "Acquire"),
        ("Page 2\nRule 1",
         "Page 2 Rule 1: writes a complete if-then rule logically consistent with their WS6 tree",
         "Deepen"),
        ("Page 2\nRule 2",
         "Page 2 Rule 2: second rule consistent with tree",
         "Deepen"),
        ("Page 2\nRules 3-6",
         "Page 2 Rules 3-6: additional rules present and consistent (score if all paths covered)",
         "Deepen"),
        ("Overall WS7", "Overall AI-CFT level for WS7 as a whole", ""),
        ("Notes WS7", "Free notes", ""),
    ],

    # ── WS10: Systematically Determine the Threshold Value ─────────────────────
    # One table task + one final blank + optional written reasoning
    "WS10 - Systematic": [
        ("Table\nMisclassification counts",
         "Table: correctly counts misclassifications for each of the 7 candidate thresholds",
         "Deepen"),
        ("Final blank\nOptimum threshold",
         "Final blank: 'The optimum threshold is: ___' — selects threshold with fewest errors",
         "Deepen"),
        ("Reasoning\nWhy optimum",
         "Written justification: explains why this threshold minimizes error (score if present)",
         "Create"),
        ("Overall WS10", "Overall AI-CFT level for WS10 as a whole", ""),
        ("Notes WS10", "Free notes", ""),
    ],

    # ── WS11: Evaluation ───────────────────────────────────────────────────────
    # Q1-Q3, Q5-Q7 are Likert scale / demographic — excluded from AI-CFT scoring
    "WS11 - Evaluation": [
        ("Q4\nCan explain DT",
         "Q4 (yes/no): Can you explain how a decision tree works? Score as Acquire if 'Yes'",
         "Acquire"),
        ("Q8a\nClassify strawberry",
         "Q8a: Uses the given decision tree to classify strawberries correctly",
         "Acquire"),
        ("Q8b\nDecision rule",
         "Q8b: Writes correct decision rule for strawberry classification (blank: ___ because ___)",
         "Acquire"),
        ("Q9\nWhat is data",
         "Q9: Open-ended (6 lines) — explains what data is and why we need it",
         "Deepen"),
        ("Q10\nDT capabilities",
         "Q10: Multi-select T/F table — marks correct statements about what DTs can do",
         "Deepen"),
        ("Q11\nStep ordering",
         "Q11: Orders the 4 steps for building a decision rule correctly (1-4)",
         "Acquire"),
        ("Q12\nWhy is DT AI",
         "Q12: Multi-select T/F table — marks correct reasons why a DT is considered AI",
         "Deepen"),
        ("Overall WS11", "Overall AI-CFT level for WS11 as a whole", ""),
        ("Notes WS11", "Free notes (Q1-Q3 and Q5-Q7 are Likert/demographic, not scored here)", ""),
    ],

    # ── WS_DT: CODAP Arbor Decision Tree Activity ──────────────────────────────
    "WS_DT - CODAP Arbor": [
        # Section A
        ("A1\nPrior belief",
         "A1: Names variable(s) predicted to determine recommendability — based on intuition, any answer ok",
         "Acquire"),
        ("A2\nData-based exploration",
         "A2: Names variable(s) based on data WITH graph/visualization evidence (not just intuition)",
         "Acquire"),
        ("A3\nMeaningful difference found",
         "A3: Yes/No response AND references actual data to support the answer",
         "Acquire"),
        ("A4\nFirst variable justification",
         "A4: Justifies first-variable choice using data evidence (not general knowledge such as 'fat is unhealthy')",
         "Deepen"),
        # Section B
        ("B1\nTree 1 — build + EMIT",
         "B1: Builds single-level tree with one variable, records EMIT accuracy result",
         "Deepen"),
        ("B2\nTree 2 — build + EMIT",
         "B2: Builds second single-level tree with a different variable, records EMIT",
         "Deepen"),
        ("B3\nTree 3 — build + EMIT",
         "B3: Builds third single-level tree with yet another variable, records EMIT",
         "Deepen"),
        ("B4\nBest tree + criteria",
         "B4: Identifies best tree among B1-B3 AND lists criteria used for comparison",
         "Deepen"),
        # Section C
        ("C1\nThreshold optimization",
         "C1: Tries different threshold values for best variable from B; saves best with EMIT",
         "Deepen"),
        ("C2\nThreshold effect",
         "C2: Explains how changing the threshold value affects decision tree performance",
         "Deepen"),
        ("C3\nBest threshold + reasoning",
         "C3: Identifies best threshold and explains how that conclusion was reached",
         "Deepen"),
        # Section D
        ("D1\nTwo-level tree — build + EMIT",
         "D1: Adds a second variable to create a two-level tree; saves with EMIT",
         "Create"),
        ("D2\nSecond level effect",
         "D2: Explains whether and how the second variable improved classification",
         "Create"),
        ("D3\nBest variable combination",
         "D3: States which variable combination produced the best result",
         "Create"),
        ("D4\nStopping criterion",
         "D4: Explains how they decided they had reached their best decision tree",
         "Create"),
        # Section E
        ("E\nMetrics: sensitivity + MCR",
         "E (intro): Records Sensitivity and MCR values from EMIT for the best tree",
         "Deepen"),
        ("E1\nMetric interpretation",
         "E1: Explains which metric matters more for this model's purpose and why",
         "Create"),
        ("E2\nError type analysis",
         "E2: Identifies whether the model made more false positives or false negatives",
         "Deepen"),
        ("E3\nGood-enough criterion",
         "E3: Reflects on how software / they would decide when a tree is good enough",
         "Create"),
        ("E4\nPerfect classification",
         "E4: Reflects on whether DTs can achieve zero error; gives a reasoned answer",
         "Create"),
        # Section F
        ("F1\nTest set application",
         "F1: Applies best tree to test dataset, evaluates with EMIT, records result",
         "Create"),
        ("F2\nTrain vs test comparison",
         "F2: Compares test vs. training performance and explains any difference",
         "Create"),
        # Section G
        ("G1\nWhat the model learned",
         "G1: Explains what the decision tree model 'learned' from the training data",
         "Create"),
        ("G2\nPersonal reflection",
         "G2: Reflects on personal learning during the tree-building process",
         "Create"),
        ("Overall WS_DT", "Overall AI-CFT level for WS_DT as a whole", ""),
        ("Notes WS_DT", "Free notes", ""),
    ],
}

# ── Helpers ────────────────────────────────────────────────────────────────────

def make_dv_adc():
    """Dropdown: Acquire / Deepen / Create / ? (unscorable)"""
    dv = DataValidation(
        type="list",
        formula1='"Acquire,Deepen,Create,?"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error       = "Choose: Acquire, Deepen, Create, or ?"
    dv.errorTitle  = "Invalid value"
    dv.prompt      = "AI-CFT level"
    dv.promptTitle = "Select level"
    return dv

def set_header_row(ws, row, labels, col_start=1):
    for i, label in enumerate(labels):
        c = ws.cell(row=row, column=col_start + i, value=label)
        c.font      = FONT_WHITE_BOLD
        c.fill      = fill(CLR["header_dark"])
        c.alignment = CENTER
        c.border    = thin_border()

def style_cell(cell, bg, fnt=None, align=None):
    cell.fill      = fill(bg)
    cell.font      = fnt or FONT_NORMAL
    cell.alignment = align or CENTER
    cell.border    = thin_border()

def col_color(header):
    h = header.lower()
    if "notes" in h or "overall" in h:
        return CLR["note"]
    return CLR["white"]

# ── Sheet: Instructions ────────────────────────────────────────────────────────

def build_instructions(wb):
    ws = wb.create_sheet("00_Instructions")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 60
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 45

    def hdr(row, text, color="1F3864"):
        ws.merge_cells(f"B{row}:E{row}")
        c = ws.cell(row=row, column=2, value=text)
        c.font      = Font(name="Calibri", bold=True, color="FFFFFF", size=12)
        c.fill      = fill(color)
        c.alignment = LEFT
        ws.row_dimensions[row].height = 24

    def row_text(row, col_b, col_c, b="", c="", d="", e="", bg=CLR["white"]):
        data = {"B": b, "C": c, "D": d, "E": e}
        for col_letter, val in data.items():
            col_num = ord(col_letter) - ord("A") + 1
            cell = ws.cell(row=row, column=col_num, value=val)
            cell.alignment = LEFT
            cell.font      = FONT_NORMAL
            cell.fill      = fill(bg)

    r = 2
    hdr(r, "AI-CFT Manual Labeling — Instructions")
    r += 2

    hdr(r, "1. Structure", "2E75B6")
    r += 1
    rows_info = [
        ("Each tab", "Corresponds to one worksheet (WS1 through WS_DT)"),
        ("Cohort 2025", "Individual student folders — one row per student (real pseudonyms)"),
        ("Cohort 2026", "Class-level PDFs — one row per student (placeholder IDs: S01, S02...)"),
        ("Overall column", "Single AI-CFT level for the worksheet as a whole"),
    ]
    for label, desc in rows_info:
        ws.cell(row=r, column=2, value=label).font = FONT_DARK_BOLD
        ws.cell(row=r, column=3, value=desc).font  = FONT_NORMAL
        ws.row_dimensions[r].height = 18
        r += 1

    r += 1
    hdr(r, "2. AI-CFT Levels", "2E75B6")
    r += 1
    levels = [
        ("Acquire",  CLR["acquire"],
         "Student recognizes, recalls, or defines a concept. One-way knowledge transfer.",
         "WS1 fill-in-the-blank terms; WS7 path matching; WS11 Q8 classification"),
        ("Deepen",   CLR["deepen"],
         "Student applies, analyzes, or compares concepts. Uses data as evidence.",
         "WS3 own threshold; WS4 B4 Pia same-fat; WS_DT A4 data justification"),
        ("Create",   CLR["create"],
         "Student designs, synthesizes, or transfers to a new context. High-order reasoning.",
         "WS10 reasoning; WS_DT D1-D4, E1, E3-E4, F1-F2, G1-G2"),
        ("?",        CLR["note"],
         "Cannot be scored: blank, illegible, off-topic, or ambiguous response.",
         "Add explanation in the Notes column"),
    ]
    for lvl, bg, desc, ex in levels:
        ws.cell(row=r, column=2, value=lvl).font   = FONT_DARK_BOLD
        ws.cell(row=r, column=2).fill              = fill(bg)
        ws.cell(row=r, column=3, value=desc).font  = FONT_NORMAL
        ws.cell(row=r, column=4, value="Examples:").font = FONT_DARK_BOLD
        ws.cell(row=r, column=5, value=ex).font    = FONT_TINY
        for col in range(2, 6):
            ws.cell(row=r, column=col).fill = fill(bg)
            ws.cell(row=r, column=col).alignment = LEFT
        ws.row_dimensions[r].height = 36
        r += 1

    r += 1
    hdr(r, "3. How to Fill", "2E75B6")
    r += 1
    steps = [
        "1 - Open the relevant tab (e.g. WS3 - Threshold Apply).",
        "2 - Find the student row (name for Cohort 2025; S01... for Cohort 2026).",
        "3 - Open the student's paper or PDF alongside this file.",
        "4 - For each task column, select a level from the dropdown (Acquire / Deepen / Create / ?).",
        "5 - Fill the Overall column with a single holistic level for that worksheet.",
        "6 - Add a short note in the Notes column if needed.",
        "7 - Save with Ctrl+S after finishing each student.",
    ]
    for step in steps:
        ws.cell(row=r, column=2, value=step).font      = FONT_NORMAL
        ws.cell(row=r, column=2).alignment             = LEFT
        ws.merge_cells(f"B{r}:E{r}")
        ws.row_dimensions[r].height = 18
        r += 1

    r += 1
    hdr(r, "4. Important Notes", "C00000")
    r += 1
    notes = [
        "The Overall level may differ from individual task levels -- that is intentional. Overall reflects holistic judgment.",
        "Re-read the same student responses in a second sitting to check your own consistency (intra-rater reliability).",
        "Replace S01... placeholders with real pseudonyms as you identify 2026 students.",
        "Do not change cell background colors manually -- the hint row colors are for orientation only.",
        "Q1-Q3 and Q5-Q7 in WS11 are Likert/demographic items and are intentionally excluded from scoring.",
    ]
    for n in notes:
        ws.cell(row=r, column=2, value="* " + n).font = Font(name="Calibri", size=10, color="C00000")
        ws.cell(row=r, column=2).alignment = LEFT
        ws.merge_cells(f"B{r}:E{r}")
        ws.row_dimensions[r].height = 20
        r += 1

    return ws

# ── Sheet: one worksheet per WS ───────────────────────────────────────────────

def build_ws_sheet(wb, sheet_name, tasks, students_2025, students_2026):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C4"   # freeze student name + cohort columns

    dv = make_dv_adc()
    ws.add_data_validation(dv)

    task_headers = [t[0] for t in tasks]
    is_scoring   = ["Notes" not in h and "Overall" not in h for h in task_headers]
    is_overall   = ["Overall" in h for h in task_headers]
    is_note      = ["Notes" in h for h in task_headers]

    # ── Row 1: sheet title ────────────────────────────────────────────────────
    ws.merge_cells(f"A1:{get_column_letter(2 + len(tasks))}1")
    title_cell = ws.cell(row=1, column=1, value=f"AI-CFT Labeling — {sheet_name}")
    title_cell.font      = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    title_cell.fill      = fill(CLR["header_dark"])
    title_cell.alignment = CENTER
    ws.row_dimensions[1].height = 28

    # ── Row 2: AI-CFT hint sub-row ────────────────────────────────────────────
    ws.cell(row=2, column=1, value="Cohort").font = FONT_DARK_BOLD
    ws.cell(row=2, column=2, value="Student").font = FONT_DARK_BOLD
    for i, task in enumerate(tasks):
        hint = task[2]
        c = ws.cell(row=2, column=3 + i, value=hint if hint else "—")
        c.font      = Font(name="Calibri", size=8, italic=True, color="595959")
        c.fill      = fill(CLR["acquire"] if hint == "Acquire" else
                           CLR["deepen"]  if hint == "Deepen"  else
                           CLR["create"]  if hint == "Create"  else CLR["note"])
        c.alignment = CENTER
    ws.row_dimensions[2].height = 16

    # ── Row 3: column headers ─────────────────────────────────────────────────
    ws.cell(row=3, column=1, value="Cohort").font = FONT_WHITE_BOLD
    ws.cell(row=3, column=1).fill = fill(CLR["header_dark"])
    ws.cell(row=3, column=1).alignment = CENTER
    ws.cell(row=3, column=2, value="Student Name").font = FONT_WHITE_BOLD
    ws.cell(row=3, column=2).fill = fill(CLR["header_dark"])
    ws.cell(row=3, column=2).alignment = CENTER
    for i, h in enumerate(task_headers):
        c = ws.cell(row=3, column=3 + i, value=h)
        c.font      = FONT_WHITE_BOLD
        c.fill      = fill(CLR["header_mid"])
        c.alignment = CENTER
        c.border    = thin_border()
    ws.row_dimensions[3].height = 36

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 10   # cohort
    ws.column_dimensions["B"].width = 18   # student name
    for i, h in enumerate(task_headers):
        col_letter = get_column_letter(3 + i)
        if "Notes" in h:
            ws.column_dimensions[col_letter].width = 40
        elif "Overall" in h:
            ws.column_dimensions[col_letter].width = 14
        else:
            ws.column_dimensions[col_letter].width = 13

    # ── Data rows ─────────────────────────────────────────────────────────────
    row = 4
    for cohort_label, students in [("2025", students_2025), ("2026", students_2026)]:
        # Cohort separator row
        ws.merge_cells(f"A{row}:{get_column_letter(2 + len(tasks))}{row}")
        sep = ws.cell(row=row, column=1, value=f"▶  Cohort {cohort_label}")
        sep.font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        sep.fill      = fill(CLR["header_mid"])
        sep.alignment = LEFT
        ws.row_dimensions[row].height = 18
        row += 1

        for idx, student in enumerate(students):
            bg = CLR["row_alt"] if idx % 2 == 0 else CLR["white"]
            ws.cell(row=row, column=1, value=cohort_label).fill = fill(bg)
            ws.cell(row=row, column=1).alignment = CENTER
            ws.cell(row=row, column=1).font = FONT_SMALL

            ws.cell(row=row, column=2, value=student).fill = fill(bg)
            ws.cell(row=row, column=2).alignment = LEFT
            ws.cell(row=row, column=2).font = FONT_NORMAL

            for i, (h, desc, hint) in enumerate(tasks):
                col = 3 + i
                col_letter = get_column_letter(col)
                c = ws.cell(row=row, column=col, value="")
                if "Notes" in h:
                    c.fill = fill(CLR["note"])
                    c.alignment = LEFT
                    c.font = FONT_SMALL
                elif "Overall" in h:
                    c.fill = fill(CLR["note"])
                    c.alignment = CENTER
                    dv.add(c)
                else:
                    c.fill = fill(bg)
                    c.alignment = CENTER
                    dv.add(c)
                c.border = thin_border()

            ws.row_dimensions[row].height = 18
            row += 1

    # ── Tooltip row at bottom ─────────────────────────────────────────────────
    row += 1
    ws.merge_cells(f"A{row}:{get_column_letter(2 + len(tasks))}{row}")
    tip = ws.cell(row=row, column=1,
                  value="Dropdown values: Acquire | Deepen | Create | ? (unscorable/blank)")
    tip.font = Font(name="Calibri", size=9, italic=True, color="595959")
    tip.alignment = LEFT

    # ── Description sub-row (row 2 hint for each task) ───────────────────────
    # Already done above via the hint row. Add task description as comment-like row:
    desc_row = row + 1
    ws.cell(desc_row, 1, "Task descriptions:").font = FONT_DARK_BOLD
    ws.cell(desc_row, 2, "").font = FONT_SMALL
    for i, (h, desc, hint) in enumerate(tasks):
        c = ws.cell(desc_row, 3 + i, desc)
        c.font = FONT_TINY
        c.alignment = CENTER
        c.fill = fill(CLR["note"])
    ws.row_dimensions[desc_row].height = 40

    return ws

# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default sheet

    build_instructions(wb)

    for sheet_name, tasks in WS_TASKS.items():
        build_ws_sheet(wb, sheet_name, tasks, STUDENTS_2025, STUDENTS_2026)

    out_path = "/Users/mrved/Desktop/DecisionTrees-AICFT-Assessor/AI_CFT_Labeling.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"Sheets: {[s.title for s in wb.worksheets]}")

if __name__ == "__main__":
    main()
